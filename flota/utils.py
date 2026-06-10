"""
Utilidades para exportación y generación de reportes
"""
import requests
import re
from datetime import datetime
from calendar import monthrange
from collections import defaultdict
from django.http import HttpResponse
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import normalizar_estado_oc, normalizar_estado_visual, Vehiculo, Mantenimiento, OrdenCompra, Proveedor, CuentaPresupuestaria, Presupuesto


MESES = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
]

def crear_estilos_excel():
    """
    Estilos reutilizables para el Excel
    """
    estilo_titulo = Font(name='Arial', size=14, bold=True)
    estilo_encabezado_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    estilo_encabezado_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    centrado = Alignment(horizontal='center', vertical='center')
    derecha = Alignment(horizontal='right', vertical='center')
    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    return {
        'titulo': estilo_titulo,
        'encabezado_font': estilo_encabezado_font,
        'encabezado_fill': estilo_encabezado_fill,
        'centrado': centrado,
        'derecha': derecha,
        'borde': borde,
    }


def consultar_oc_mercado_publico(codigo_oc):
    """
    Consulta la API de Mercado Público y retorna un diccionario con datos limpios
    """
    try:
        ticket = getattr(settings, 'MERCADO_PUBLICO_TICKET', None)

        if not ticket or ticket == 'BLABLABLA':
            return {'error': 'Ticket no configurado.'}
        
        url = f"https://api.mercadopublico.cl/servicios/v1/publico/ordenesdecompra.json?codigo={codigo_oc}&ticket={ticket}"
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
        }
        
        response = requests.get(url, headers=headers, timeout=30, verify=True)
        response.raise_for_status()
        
        data = response.json()

        # PARA DEPURACIÓN
        print("=== RESPUESTA API ===")
        print(data)
        print("====================")
        
        if data.get('Cantidad', 0) == 0:
            return {'error': f'No se encontró la orden de compra {codigo_oc}.'}

        if 'Listado' not in data or not data['Listado']:
            return {'error': 'La API no devolvió datos en el formato esperado.'}

        oc_data = data['Listado'][0]

        # Extraer items
        items_str = ""
        if 'Listado' in oc_data.get('Items', {}):
            items = []
            lista = oc_data['Items']['Listado']
            if isinstance(lista, dict):
                lista = [lista]
                
            for item in lista:
                nombre = item.get('Producto', '')
                espec = item.get('EspecificacionComprador', '')
                items.append(f"{nombre} - {espec}")
            
            items_str = "\n".join(items)

        # Datos básicos
        estado_original = oc_data.get('Estado', 'Emitida')
        
        estado_normalizado = normalizar_estado_oc(estado_original)
        estado_visual = normalizar_estado_visual(estado_normalizado)

        # Datos básicos
        info_limpia = {
            'codigo': oc_data.get('Codigo', codigo_oc),
            'estado_original': estado_original,
            'fecha_emision': oc_data.get('Fechas', {}).get('FechaCreacion', '').split('T')[0],
            'descripcion': oc_data.get('Descripcion', oc_data.get('Nombre', f'Orden de compra {codigo_oc}'))[:500],
            'monto_neto': oc_data.get('TotalNeto', 0),
            'monto_total': oc_data.get('Total', 0),
            'impuestos': oc_data.get('Impuestos', 0),
            'proveedor_rut': oc_data.get('Proveedor', {}).get('RutSucursal', ''),
            'proveedor_nombre': oc_data.get('Proveedor', {}).get('Nombre', 'Proveedor Desconocido'),
            'id_licitacion': oc_data.get('CodigoLicitacion', ''),
            'items_str': items_str
        }

        # Extraer tipo de adquisición
        tipo_adquisicion = 'Convenio Marco'  # valor por defecto
        if 'TipoAdquisicion' in oc_data:
            raw_tipo = oc_data['TipoAdquisicion']
            # Normalizar a las opciones del modelo
            if 'CONVENIO' in raw_tipo.upper():
                tipo_adquisicion = 'Convenio Marco'
            elif 'LICITACIÓN' in raw_tipo.upper() or 'LICITACION' in raw_tipo.upper():
                tipo_adquisicion = 'Licitación Pública'
            elif 'TRATO DIRECTO' in raw_tipo.upper():
                tipo_adquisicion = 'Trato Directo'
            elif 'COMPRA ÁGIL' in raw_tipo.upper() or 'COMPRA AGIL' in raw_tipo.upper():
                tipo_adquisicion = 'Compra Ágil'
            else:
                tipo_adquisicion = raw_tipo  # mantener original si no coincide

        info_limpia['tipo_adquisicion'] = tipo_adquisicion
        
        # ========== EXTRACCIÓN DE DATOS CLAVE ==========
        texto_completo = f"{info_limpia['descripcion']} {items_str}".upper()
        
        # 1. Buscar CÓDIGO PRESUPUESTARIO (22.06.002.002)
        import re
        codigo_match = re.search(r'(\d{2}\.\d{2}\.\d{3}\.\d{3})', texto_completo)
        if codigo_match:
            info_limpia['codigo_presupuestario'] = codigo_match.group(1)
        else:
            # Intentar con formato 22-06-002
            codigo_match = re.search(r'(\d{2}-\d{2}-\d{3})', texto_completo)
            if codigo_match:
                info_limpia['codigo_presupuestario'] = codigo_match.group(1).replace('-', '.')
            else:
                info_limpia['codigo_presupuestario'] = None

        # 2. Buscar PATENTES con patrones mejorados
        patentes_encontradas = []

        # Nuevos patrones para capturar formatos específicos
        patrones_patentes = [
            r'\b(HR\.PG-25)\b',  # Patente específica de este caso
            r'\b([A-Z]{2}\.[A-Z]{2}-\d{2})\b',  # Formato XX.XX-XX (HR.PG-25)
            r'\b([A-Z]{2}-[A-Z]{2}-\d{2})\b',   # Formato XX-XX-XX (LX-FG-16)
            r'\b([A-Z]{2}\.[A-Z]{2}\d{2})\b',   # Formato XX.XXXX (HR.PG25)
            r'\b([A-Z]{4}\d{2})\b',             # Formato XXXXNN (HRPG25)
            r'\b([A-Z]{2}\d{4})\b',             # Formato XXNNNN
            r'Patente Vehículo\s+([A-Z0-9\.\-]+)',  # Buscar específicamente después de "Patente Vehículo"
        ]

        for patron in patrones_patentes:
            matches = re.findall(patron, texto_completo, re.IGNORECASE)
            if matches:
                print(f"Patrón '{patron}' encontró: {matches}")
                patentes_encontradas.extend(matches)

        # También buscar en todo el texto sin patrones específicos
        # Buscar cualquier cosa que parezca patente chilena
        if not patentes_encontradas:
            # Patrón general para patentes chilenas
            patron_general = r'\b([A-Z]{2}[\.\-]?[A-Z]{2}[\.\-]?\d{2,3})\b'
            matches_general = re.findall(patron_general, texto_completo, re.IGNORECASE)
            if matches_general:
                print(f"Patrón general encontró: {matches_general}")
                patentes_encontradas.extend(matches_general)

        # Limpiar y normalizar
        patentes_limpias = []
        for patente in set(patentes_encontradas):
            # Convertir a mayúsculas
            patente = patente.upper()
            
            # Si ya tiene formato con punto y guión (HR.PG-25), mantenerlo
            if re.match(r'^[A-Z]{2}\.[A-Z]{2}-\d{2}$', patente):
                patentes_limpias.append(patente)
            # Si tiene formato con guión (LX-FG-16), mantenerlo
            elif re.match(r'^[A-Z]{2}-[A-Z]{2}-\d{2}$', patente):
                patentes_limpias.append(patente)
            # Si tiene formato compacto (HRPG25), formatearlo
            elif re.match(r'^[A-Z]{4}\d{2}$', patente):
                patente_formateada = f"{patente[:2]}.{patente[2:4]}-{patente[4:]}"
                patentes_limpias.append(patente_formateada)
            else:
                # Mantener cualquier otro formato
                patentes_limpias.append(patente)

        info_limpia['patentes_posibles'] = patentes_limpias
        
        return info_limpia

    except Exception as e:
        return {'error': f'Error: {str(e)}'}

def exportar_reporte_excel(titulo, datos, columnas, nombre_archivo=None):
    """
    Genera un archivo Excel a partir de datos
    
    Args:
        titulo: Título del reporte
        datos: Lista de diccionarios con los datos
        columnas: Lista de tuplas (nombre_columna, clave_dato, formato)
        nombre_archivo: Nombre del archivo (opcional)
    
    Returns:
        HttpResponse con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    estilos = crear_estilos_excel()
    
    # Título
    ws.merge_cells('A1:{}1'.format(get_column_letter(len(columnas))))
    celda_titulo = ws['A1']
    celda_titulo.value = titulo
    celda_titulo.font = estilos['titulo']
    celda_titulo.alignment = estilos['centrado']
    ws.row_dimensions[1].height = 25
    
    # Fecha de generación
    ws.merge_cells('A2:{}2'.format(get_column_letter(len(columnas))))
    celda_fecha = ws['A2']
    celda_fecha.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    celda_fecha.alignment = estilos['centrado']
    ws.row_dimensions[2].height = 20
    
    # Encabezados
    fila_encabezado = 3
    for idx, (nombre_col, _, _) in enumerate(columnas, start=1):
        celda = ws.cell(row=fila_encabezado, column=idx)
        celda.value = nombre_col
        celda.font = estilos['encabezado_font']
        celda.fill = estilos['encabezado_fill']
        celda.alignment = estilos['centrado']
        celda.border = estilos['borde']
    
    ws.row_dimensions[fila_encabezado].height = 20
    
    # Datos
    fila_actual = fila_encabezado + 1
    for item in datos:
        for idx, (_, clave_dato, formato) in enumerate(columnas, start=1):
            celda = ws.cell(row=fila_actual, column=idx)
            valor = item.get(clave_dato, '')
            
            if formato == 'moneda':
                celda.value = float(valor) if valor else 0
                celda.number_format = '$#,##0'
            elif formato == 'decimal':
                celda.value = float(valor) if valor else 0
                celda.number_format = '#,##0.00'
            elif formato == 'entero':
                celda.value = int(valor) if valor else 0
            elif formato == 'fecha':
                if valor:
                    if hasattr(valor, 'strftime'):
                        celda.value = valor.strftime('%d/%m/%Y')
                    else:
                        celda.value = str(valor)
            else:
                celda.value = str(valor) if valor else ''
            
            celda.border = estilos['borde']
            if formato in ['moneda', 'decimal', 'entero']:
                celda.alignment = estilos['derecha']
            else:
                celda.alignment = Alignment(horizontal='left', vertical='center')
        
        fila_actual += 1
    
    # Ajustar ancho de columnas
    for idx, (nombre_col, _, _) in enumerate(columnas, start=1):
        col_letter = get_column_letter(idx)
        # Ancho mínimo basado en el nombre de la columna
        ancho = max(len(nombre_col) + 2, 12)
        ws.column_dimensions[col_letter].width = ancho
    
    # Preparar respuesta
    if not nombre_archivo:
        nombre_archivo = f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    if not nombre_archivo.endswith('.xlsx'):
        nombre_archivo += '.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    return response


def aplicar_estilos_cabecera(ws, fila, columnas, estilos):
    """
    Aplica estilos a una fila de encabezados
    """
    for idx, nombre in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=idx)
        celda.value = nombre
        celda.font = estilos['encabezado_font']
        celda.fill = estilos['encabezado_fill']
        celda.alignment = estilos['centrado']
        celda.border = estilos['borde']
        ws.row_dimensions[fila].height = 20


def exportar_planilla_mantenimientos_excel(anio=None):
    """
    Exporta un Excel con 4 hojas en el formato oficial del hospital:
    - CATASTRO Y PLANIFICACIÓN MP
    - SEGUIMIENTO GASTO
    - Hoja1 (resumen neumáticos)
    - TABLA CÁLCULO (indicadores)
    """
    if not anio:
        anio = datetime.now().year
    else:
        anio = int(anio)

    wb = Workbook()
    estilos = crear_estilos_excel()

    # ========== 1. CATASTRO Y PLANIFICACIÓN MP ==========
    ws1 = wb.active
    ws1.title = "CATASTRO Y PLANIFICACIÓN MP"

    # Título
    ws1.merge_cells('A1:{}1'.format(get_column_letter(52)))  # 52 columnas
    ws1['A1'].value = f"CATASTRO Y PLANIFICACIÓN MP - AÑO {anio}"
    ws1['A1'].font = estilos['titulo']
    ws1['A1'].alignment = estilos['centrado']

    # Encabezados (52 columnas, incluyendo las mensuales)
    encabezados_fijos = [
        'REGIÓN', 'ESTABLECIMIENTO', 'TIPO CARROCERÍA', 'TIPO AMBULANCIA',
        'CLASE DE AMBULANCIA', 'SAMU (SI / NO)', 'FUNCIÓN', 'MARCA', 'MODELO',
        'N° PATENTE', 'N° MOTOR', 'KILOMETRAJE', 'ESTADO SITUACIÓN',
        'ESTADO DE CONSERVACIÓN (BUENO / REGULAR / MALO / BAJA)',
        'AÑO ADQUISICIÓN', 'VIDA ÚTIL', 'VIDA ÚTIL RESIDUAL',
        'CRÍTICO / NO CRÍTICO', 'EN GARANTÍA (SI / NO)', 'AÑO VENCIMIENTO GARANTÍA',
        'BAJO PLAN DE MANTENIMIENTO (SI / NO)', 'AÑO PLAN DE MANTENIMIENTO',
        'MANTENIMIENTO INTERNO O MANTENIMIENTO EXTERNO O CONTRATO',
        'NOMBRE DE PROVEEDOR O MANTENIMIENTO INTERNO', 'ID CONVENIO DE MANTENIMIENTO',
        'COSTO ANUAL DE MANTENIMIENTO SEGÚN CONVENIO / PRECIO DE REFERENCIA MANTENIMIENTO ANUAL',
        'KILOMETRAJE RECORRIDO [KM] ACTUALIZADO', 'FRECUENCIA ANUAL DE MANTENIMIENTO'
    ]
    # Columnas mensuales: para cada mes, dos columnas: TIPO y ESTADO
    columnas_mensuales = []
    for mes in MESES:
        columnas_mensuales.append(f'TIPO {mes}')
        columnas_mensuales.append(f'ESTADO {mes}')
    encabezados = encabezados_fijos + columnas_mensuales

    fila_enc = 2
    for idx, header in enumerate(encabezados, start=1):
        celda = ws1.cell(row=fila_enc, column=idx)
        celda.value = header
        celda.font = estilos['encabezado_font']
        celda.fill = estilos['encabezado_fill']
        celda.alignment = estilos['centrado']
        celda.border = estilos['borde']
    ws1.row_dimensions[fila_enc].height = 30

    # Datos: obtener vehículos y conceptos
    vehiculos = Vehiculo.objects.all().order_by('patente')
    # Pre-cargar relaciones para eficiencia
    mantenimientos = Mantenimiento.objects.filter(fecha_ingreso__year=anio).select_related('vehiculo', 'proveedor')
    ordenes_compra = OrdenCompra.objects.filter(fecha_emision__year=anio, monto_total__gt=0).select_related('vehiculo', 'proveedor', 'cuenta_presupuestaria')

    # Diccionarios auxiliares
    mant_por_vehiculo = defaultdict(list)
    for m in mantenimientos:
        mant_por_vehiculo[m.vehiculo_id].append(m)

    oc_por_vehiculo = defaultdict(list)
    for oc in ordenes_compra:
        if oc.vehiculo_id:
            oc_por_vehiculo[oc.vehiculo_id].append(oc)

    # Para cada vehículo, identificar conceptos (principal + adicionales)
    fila_actual = fila_enc + 1

    for v in vehiculos:
        # Datos fijos del vehículo
        region = "LOS LAGOS"
        establecimiento = v.establecimiento or "HOSPITAL DE RIO NEGRO"
        tipo_carroceria = v.get_tipo_carroceria_display()
        tipo_ambulancia = "EMERGENCIA BASICA" if v.tipo_carroceria == 'Ambulancia' else ""
        clase_ambulancia = v.clase_ambulancia or ""
        samu = "SI" if v.es_samu else "NO"
        funcion = "TRANSPORTE DE EMERGENCIA" if v.tipo_carroceria == 'Ambulancia' else "TRANSPORTE DE FUNCIONARIOS"
        marca = v.marca
        modelo = v.modelo
        patente = v.patente
        nro_motor = v.nro_motor
        kilometraje = v.kilometraje_actual
        estado_situacion = v.get_tipo_propiedad_display()
        estado_conservacion = "BUENO"  # Por defecto, sin campo en BD
        anio_adq = v.anio_adquisicion
        vida_util = v.vida_util
        vida_util_residual = max(0, vida_util - (anio - anio_adq))
        criticidad = v.get_criticidad_display()
        en_garantia = "NO"
        anio_vencimiento_garantia = ""
        bajo_plan = "SI" if mant_por_vehiculo.get(v.id) else "NO"
        anio_plan = anio if bajo_plan == "SI" else ""
        tipo_mant_contrato = "CONTRATO"
        nombre_proveedor = ""
        id_convenio = ""
        costo_anual_convenio = 0
        proveedores_mp = set()
        for m in mant_por_vehiculo.get(v.id, []):
            if m.tipo_mantencion == 'Preventivo' and m.proveedor:
                proveedores_mp.add(m.proveedor_id)
        if proveedores_mp:
            # Tomar el primero (podría afinarse)
            proveedor_id = next(iter(proveedores_mp))
            proveedor = Proveedor.objects.get(id=proveedor_id)
            nombre_proveedor = proveedor.nombre_fantasia
            # Buscar id_convenio en las OC asociadas
            for oc in oc_por_vehiculo.get(v.id, []):
                if oc.proveedor_id == proveedor_id and oc.id_licitacion:
                    id_convenio = oc.id_licitacion
                    break
            # Sumar todas las OC de ese proveedor para este vehículo en el año
            costo_anual_convenio = sum(oc.monto_total for oc in oc_por_vehiculo.get(v.id, []) if oc.proveedor_id == proveedor_id)
        frecuencia_anual = len([m for m in mant_por_vehiculo.get(v.id, []) if m.tipo_mantencion == 'Preventivo'])

        # Datos de mantenimientos por mes (12 meses)
        # Inicializar diccionario: mes -> (tipo, estado)
        datos_mensuales = {mes: ('', '') for mes in range(1, 13)}
        for m in mant_por_vehiculo.get(v.id, []):
            mes = m.fecha_ingreso.month
            # Tipo: puede ser la descripción del trabajo o un resumen
            tipo = m.descripcion_trabajo[:50] if m.descripcion_trabajo else m.get_tipo_mantencion_display()
            estado = '√' if m.estado == 'Finalizado' else ('ꓣ' if m.estado == 'Programado' else 'X')
            datos_mensuales[mes] = (tipo, estado)

        # Ahora construir la fila principal del vehículo
        fila_base = [
            region, establecimiento, tipo_carroceria, tipo_ambulancia,
            clase_ambulancia, samu, funcion, marca, modelo, patente, nro_motor,
            kilometraje, estado_situacion, estado_conservacion, anio_adq,
            vida_util, vida_util_residual, criticidad, en_garantia,
            anio_vencimiento_garantia, bajo_plan, anio_plan,
            tipo_mant_contrato, nombre_proveedor, id_convenio,
            costo_anual_convenio, kilometraje, frecuencia_anual
        ]
        # Agregar datos mensuales (tipo y estado para cada mes)
        for mes in range(1, 13):
            tipo, estado = datos_mensuales[mes]
            fila_base.append(tipo)
            fila_base.append(estado)

        # Escribir fila
        for idx, valor in enumerate(fila_base, start=1):
            celda = ws1.cell(row=fila_actual, column=idx)
            celda.value = valor
            celda.border = estilos['borde']
            if idx in (12, 27):  # kilometraje
                celda.alignment = estilos['derecha']
            elif 26 <= idx <= 28:  # costos y frecuencia
                celda.alignment = estilos['derecha']
                if isinstance(valor, (int, float)):
                    celda.number_format = '#,##0'
        fila_actual += 1

        # ===== Conceptos adicionales (Neumáticos, Cabina sanitaria, etc.) =====
        # Buscamos órdenes de compra y mantenimientos que contengan palabras clave
        conceptos = {}
        for oc in oc_por_vehiculo.get(v.id, []):
            desc = oc.descripcion.upper() if oc.descripcion else ''
            if 'NEUMÁTICO' in desc:
                conceptos['Neumáticos'] = oc
            if 'CABINA SANITARIA' in desc:
                conceptos['Cabina sanitaria'] = oc
        # Además, detectar si hay mantenimientos con "neumático" en descripción
        for m in mant_por_vehiculo.get(v.id, []):
            desc = m.descripcion_trabajo.upper() if m.descripcion_trabajo else ''
            if 'NEUMÁTICO' in desc:
                conceptos['Neumáticos'] = m  # sobreescribir con el último

        for concepto, obj in conceptos.items():
            fila_concepto = [
                region, establecimiento, tipo_carroceria, '', '', '', '', '', '', patente, '',
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''
            ]
            # Completar las 28 primeras columnas vacías (excepto patente y algunos)
            # Luego añadir datos mensuales del concepto
            # Para simplificar, reutilizamos los mismos datos mensuales del vehículo pero con estado del concepto
            datos_concepto = {mes: ('', '') for mes in range(1, 13)}
            if hasattr(obj, 'fecha_emision'):  # es orden de compra
                mes = obj.fecha_emision.month
                datos_concepto[mes] = (f"{obj.monto_total}", '√')
            elif hasattr(obj, 'fecha_ingreso'):
                mes = obj.fecha_ingreso.month
                datos_concepto[mes] = (f"${obj.costo_total_real}", '√')

            for mes in range(1, 13):
                tipo, estado = datos_concepto[mes]
                fila_concepto.append(tipo)
                fila_concepto.append(estado)
            # Escribir fila del concepto
            for idx, valor in enumerate(fila_concepto, start=1):
                celda = ws1.cell(row=fila_actual, column=idx)
                celda.value = valor
                celda.border = estilos['borde']
            fila_actual += 1

    # Ajustar anchos de columnas (ejemplo básico)
    for i in range(1, 53):
        ws1.column_dimensions[get_column_letter(i)].width = 12

    # ========== 2. SEGUIMIENTO GASTO ==========
    ws2 = wb.create_sheet("SEGUIMIENTO GASTO")
    # Título
    ws2.merge_cells('A1:AI1')
    ws2['A1'].value = f"SEGUIMIENTO GASTO - AÑO {anio}"
    ws2['A1'].font = estilos['titulo']
    ws2['A1'].alignment = estilos['centrado']

    # --- Subtabla: Convenios de Mantenimiento (MP) ---
    ws2['A3'] = "CONVENIOS DE MANTENIMIENTO (MP)"
    ws2['A3'].font = Font(bold=True, size=12)

    encabezados_mp = [
        'Nº', 'SERVICIO DE SALUD', 'ESTABLECIMIENTO', 'NOMBRE CONVENIO DE MANTENIMIENTO',
        'Nº RESOLUCIÓN / ID / ORDEN DE COMPRA DEL CONVENIO', 'FECHA RESOLUCIÓN',
        'FECHA DE EXPIRACION CONVENIO', 'MONTO ANUAL ($)', 'SUBASIGNACIÓN SIGFE'
    ]
    # Agregar columnas mensuales: N° ORDEN DE COMPRA y MONTO ($) para cada mes
    for mes in MESES:
        encabezados_mp.append(f'N° OC {mes}')
        encabezados_mp.append(f'MONTO {mes}')
    encabezados_mp.append('TOTAL')

    fila_enc_mp = 4
    for idx, header in enumerate(encabezados_mp, start=1):
        celda = ws2.cell(row=fila_enc_mp, column=idx)
        celda.value = header
        celda.font = estilos['encabezado_font']
        celda.fill = estilos['encabezado_fill']
        celda.alignment = estilos['centrado']
        celda.border = estilos['borde']
    ws2.row_dimensions[fila_enc_mp].height = 25

    # Agrupar órdenes de compra por proveedor de convenio (es_taller=True y proveedor_base o similar)
    # Se usan los proveedores con es_taller=True y que aparezcan en OC con cuenta 22.06.002.001 o .003 (preventivo)
    proveedores_mp = Proveedor.objects.filter(es_taller=True, activo=True)
    convenios = {}
    for p in proveedores_mp:
        # Obtener todas las OC de este proveedor en el año con cuenta preventivo
        ocs = OrdenCompra.objects.filter(
            proveedor=p,
            fecha_emision__year=anio,
            cuenta_presupuestaria__codigo__in=['22.06.002.001', '22.06.002.003']
        )
        if not ocs:
            continue
        total_anual = sum(oc.monto_total for oc in ocs)
        # Determinar nombre del convenio (puede ser el nombre del proveedor)
        nombre_convenio = p.nombre_fantasia
        id_convenio = ocs.first().id_licitacion if ocs.first().id_licitacion else ''
        # Agrupar por mes
        monthly = {mes: {'nro_oc': '', 'monto': 0} for mes in range(1, 13)}
        for oc in ocs:
            mes = oc.fecha_emision.month
            monthly[mes]['nro_oc'] = oc.nro_oc
            monthly[mes]['monto'] += oc.monto_total
        convenios[p.id] = {
            'nombre': nombre_convenio, 'id': id_convenio,
            'total': total_anual, 'monthly': monthly,
            'cuenta': ocs.first().cuenta_presupuestaria.codigo if ocs.first().cuenta_presupuestaria else ''
        }

    fila_datos_mp = fila_enc_mp + 1
    for idx, (proveedor_id, data) in enumerate(convenios.items(), start=1):
        fila = [
            idx, 'OSORNO', 'HOSPITAL DE RIO NEGRO', data['nombre'],
            data['id'], '', '', data['total'], data['cuenta']
        ]
        for mes in range(1, 13):
            fila.append(data['monthly'][mes]['nro_oc'])
            fila.append(data['monthly'][mes]['monto'])
        fila.append(data['total'])  # total final
        for col, valor in enumerate(fila, start=1):
            celda = ws2.cell(row=fila_datos_mp, column=col)
            celda.value = valor
            celda.border = estilos['borde']
            if col >= 10 and col % 2 == 0:  # montos
                if isinstance(valor, (int, float)):
                    celda.number_format = '$#,##0'
                    celda.alignment = estilos['derecha']
        fila_datos_mp += 1

    # --- Subtabla: Mantenimiento Correctivo (MC) ---
    fila_offset = fila_datos_mp + 2
    ws2.cell(row=fila_offset, column=1, value="MANTENIMIENTO CORRECTIVO / ADQUISICIÓN DE REPUESTOS / ADQUISICIÓN DE INSUMOS / ADQUISICIÓN DE ACCESORIOS").font = Font(bold=True, size=12)
    fila_offset += 1

    encabezados_mc = [
        'Nº', 'SERVICIO DE SALUD', 'ESTABLECIMIENTO', 'DESCRIPCIÓN DE LA COMPRA',
        'TIPO DE GASTO CORRECTIVO', 'MONTO'
    ]
    for mes in MESES:
        encabezados_mc.append(f'N° OC {mes}')
        encabezados_mc.append(f'MONTO {mes}')
    encabezados_mc.append('TOTAL')

    fila_enc_mc = fila_offset
    for idx, header in enumerate(encabezados_mc, start=1):
        celda = ws2.cell(row=fila_enc_mc, column=idx)
        celda.value = header
        celda.font = estilos['encabezado_font']
        celda.fill = estilos['encabezado_fill']
        celda.alignment = estilos['centrado']
        celda.border = estilos['borde']
    ws2.row_dimensions[fila_enc_mc].height = 25

    # Obtener todas las OC con cuenta correctivo (22.06.002.002 o .004) y que tengan descripción
    ocs_correctivo = OrdenCompra.objects.filter(
        fecha_emision__year=anio,
        cuenta_presupuestaria__codigo__in=['22.06.002.002', '22.06.002.004']
    ).order_by('fecha_emision')

    fila_datos_mc = fila_enc_mc + 1
    for idx, oc in enumerate(ocs_correctivo, start=1):
        desc = oc.descripcion or "Sin descripción"
        tipo_gasto = "MANTENIMIENTO CORRECTIVO"
        monto = oc.monto_total
        # Agrupar por mes (para esta OC, solo un mes)
        monthly = {mes: {'nro_oc': '', 'monto': 0} for mes in range(1, 13)}
        mes = oc.fecha_emision.month
        monthly[mes]['nro_oc'] = oc.nro_oc
        monthly[mes]['monto'] = monto
        fila = [idx, 'OSORNO', 'HOSPITAL RIO NEGRO', desc, tipo_gasto, monto]
        for m in range(1, 13):
            fila.append(monthly[m]['nro_oc'])
            fila.append(monthly[m]['monto'])
        fila.append(monto)
        for col, valor in enumerate(fila, start=1):
            celda = ws2.cell(row=fila_datos_mc, column=col)
            celda.value = valor
            celda.border = estilos['borde']
            if (col >= 6 and col <= len(fila)) and (col - 5) % 2 == 0:  # montos
                if isinstance(valor, (int, float)):
                    celda.number_format = '$#,##0'
                    celda.alignment = estilos['derecha']
        fila_datos_mc += 1

    # Ajustar anchos hoja2
    for col in range(1, 4+12*2+2):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # ========== 3. Hoja1 (resumen neumáticos) ==========
    ws3 = wb.create_sheet("Hoja1")
    ws3['A1'] = "Mantenimiento móviles"
    ws3['A1'].font = estilos['titulo']
    encabezados_neumaticos = ['PATENTE', 'INSUMO/MATERIA', '$ ESTIMADO ANUAL', '$ CONSUMO ACTUAL', 'REMANENTE', 'REDUCCIÓN']
    fila_enc_neu = 3
    for idx, header in enumerate(encabezados_neumaticos, start=1):
        celda = ws3.cell(row=fila_enc_neu, column=idx)
        celda.value = header
        celda.font = estilos['encabezado_font']
        celda.fill = estilos['encabezado_fill']
        celda.alignment = estilos['centrado']
        celda.border = estilos['borde']

    # Agrupar gastos en neumáticos por vehículo
    neumaticos_por_vehiculo = defaultdict(int)
    for oc in ocs_correctivo:
        if oc.descripcion and 'NEUMÁTICO' in oc.descripcion.upper():
            if oc.vehiculo_id:
                neumaticos_por_vehiculo[oc.vehiculo.patente] += oc.monto_total
    # También considerar OC de preventivo que mencionen neumáticos
    for oc in ordenes_compra:
        if oc.descripcion and 'NEUMÁTICO' in oc.descripcion.upper() and oc.vehiculo_id:
            neumaticos_por_vehiculo[oc.vehiculo.patente] += oc.monto_total

    fila_neu = fila_enc_neu + 1
    for patente, gasto in neumaticos_por_vehiculo.items():
        vehiculo = Vehiculo.objects.filter(patente=patente).first()
        if not vehiculo:
            continue
        estimado = 1880000 if vehiculo.tipo_carroceria == 'Ambulancia' else 940000
        remanente = estimado - gasto
        reduccion = 0 if remanente >= 0 else -remanente
        fila = [patente, 'NEUMÁTICOS', estimado, gasto, max(0, remanente), reduccion]
        for col, valor in enumerate(fila, start=1):
            celda = ws3.cell(row=fila_neu, column=col)
            celda.value = valor
            celda.border = estilos['borde']
            if col >= 3:
                celda.number_format = '$#,##0'
                celda.alignment = estilos['derecha']
        fila_neu += 1

    # ========== 4. TABLA CÁLCULO ==========
    ws4 = wb.create_sheet("TABLA CÁLCULO")
    ws4['A1'] = "INDICADORES DE MANTENIMIENTO PREVENTIVO"
    ws4['A1'].font = estilos['titulo']

    # Encabezados fila 3
    encabezados_calculo = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic', 'TOTAL']
    for col, mes in enumerate(encabezados_calculo, start=1):
        celda = ws4.cell(row=3, column=col)
        celda.value = mes
        celda.font = estilos['encabezado_font']
        celda.fill = estilos['encabezado_fill']
        celda.alignment = estilos['centrado']
        celda.border = estilos['borde']

    # Calcular métricas
    # N° MP PROGRAMADAS (estados no finalizados: Programado, En taller, Esperando repuestos)
    programadas_por_mes = {m: 0 for m in range(1,13)}
    ejecutadas_por_mes = {m: 0 for m in range(1,13)}
    reprogramadas_por_mes = {m: 0 for m in range(1,13)}
    for m in mantenimientos:
        mes = m.fecha_ingreso.month
        if m.estado in ['Programado', 'En taller', 'Esperando repuestos']:
            programadas_por_mes[mes] += 1
        if m.estado == 'Finalizado':
            ejecutadas_por_mes[mes] += 1
        # Reprogramadas: si fecha_programada existe y es diferente de fecha_ingreso
        if m.fecha_programada and m.fecha_ingreso and m.fecha_programada != m.fecha_ingreso:
            reprogramadas_por_mes[mes] += 1

    # Totales
    total_programadas = sum(programadas_por_mes.values())
    total_ejecutadas = sum(ejecutadas_por_mes.values())
    total_reprogramadas = sum(reprogramadas_por_mes.values())

    # Escribir filas
    filas_calculo = [
        ('N° MP PROGRAMADAS', programadas_por_mes, total_programadas),
        ('N° MP EJECUTADAS', ejecutadas_por_mes, total_ejecutadas),
        ('N° MP REPROGRAMADAS', reprogramadas_por_mes, total_reprogramadas),
    ]
    fila_actual_calc = 4
    for nombre, datos_mes, total in filas_calculo:
        celda = ws4.cell(row=fila_actual_calc, column=1)
        celda.value = nombre
        celda.border = estilos['borde']
        for mes in range(1, 13):
            celda = ws4.cell(row=fila_actual_calc, column=mes+1)
            celda.value = datos_mes[mes]
            celda.border = estilos['borde']
        celda = ws4.cell(row=fila_actual_calc, column=14)
        celda.value = total
        celda.border = estilos['borde']
        fila_actual_calc += 1

    # Porcentajes de ejecución
    fila_actual_calc += 1
    ws4.cell(row=fila_actual_calc, column=1, value="% EJECUCIÓN MP TOTAL MENSUAL")
    for mes in range(1, 13):
        prog = programadas_por_mes[mes]
        ejec = ejecutadas_por_mes[mes]
        porc = (ejec / prog * 100) if prog > 0 else 0
        celda = ws4.cell(row=fila_actual_calc, column=mes+1)
        celda.value = porc
        celda.number_format = '0.00'
        celda.border = estilos['borde']
    ws4.cell(row=fila_actual_calc, column=14, value=(total_ejecutadas/total_programadas*100 if total_programadas else 0)).number_format = '0.00'

    # Gastos programados y ejecutados (MP y MC)
    # Gasto programado MP: suma de montos de OC de preventivo
    # Gasto ejecutado MP: suma de costos de mantenimientos preventivos finalizados
    gasto_programado_mp = sum(oc.monto_total for oc in ordenes_compra if oc.cuenta_presupuestaria and oc.cuenta_presupuestaria.codigo in ['22.06.002.001','22.06.002.003'])
    gasto_ejecutado_mp = sum(m.costo_total_real for m in mantenimientos if m.tipo_mantencion == 'Preventivo' and m.estado == 'Finalizado')
    gasto_programado_mc = sum(oc.monto_total for oc in ordenes_compra if oc.cuenta_presupuestaria and oc.cuenta_presupuestaria.codigo in ['22.06.002.002','22.06.002.004'])
    gasto_ejecutado_mc = sum(m.costo_total_real for m in mantenimientos if m.tipo_mantencion == 'Correctivo' and m.estado == 'Finalizado')

    fila_actual_calc += 2
    ws4.cell(row=fila_actual_calc, column=1, value="GASTO PROGRAMADO MP").font = Font(bold=True)
    ws4.cell(row=fila_actual_calc, column=2, value=gasto_programado_mp).number_format = '$#,##0'
    fila_actual_calc += 1
    ws4.cell(row=fila_actual_calc, column=1, value="EJECUCIÓN GASTO MP").font = Font(bold=True)
    ws4.cell(row=fila_actual_calc, column=2, value=gasto_ejecutado_mp).number_format = '$#,##0'
    fila_actual_calc += 1
    ws4.cell(row=fila_actual_calc, column=1, value="% EJECUCIÓN DEL GASTO MP").font = Font(bold=True)
    ws4.cell(row=fila_actual_calc, column=2, value=(gasto_ejecutado_mp/gasto_programado_mp*100 if gasto_programado_mp else 0)).number_format = '0.00'
    fila_actual_calc += 2
    ws4.cell(row=fila_actual_calc, column=1, value="GASTO PROGRAMADO MC").font = Font(bold=True)
    ws4.cell(row=fila_actual_calc, column=2, value=gasto_programado_mc).number_format = '$#,##0'
    fila_actual_calc += 1
    ws4.cell(row=fila_actual_calc, column=1, value="EJECUCIÓN GASTO MC").font = Font(bold=True)
    ws4.cell(row=fila_actual_calc, column=2, value=gasto_ejecutado_mc).number_format = '$#,##0'
    fila_actual_calc += 1
    ws4.cell(row=fila_actual_calc, column=1, value="% EJECUCIÓN DEL GASTO MC").font = Font(bold=True)
    ws4.cell(row=fila_actual_calc, column=2, value=(gasto_ejecutado_mc/gasto_programado_mc*100 if gasto_programado_mc else 0)).number_format = '0.00'

    # Ajustar anchos
    for col in range(1, 15):
        ws4.column_dimensions[get_column_letter(col)].width = 15

    # Guardar respuesta
    nombre_archivo = f'PLANILLA_MANTENIMIENTO_VEHICULOS_{anio}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response
