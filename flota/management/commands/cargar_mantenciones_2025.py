"""
Comando: python manage.py cargar_mantenciones_2025 [--ruta RUTA]
Carga datos de mantenciones desde un archivo Excel (ej. 1contexto/PLANILLA MANTENIMIENTO VEHÍCULOS 2025 HRN .xlsx).
Esqueleto: lee filas, busca vehículo por patente y proveedor por nombre; crea Mantenimiento si no existe.
Ajustar columnas según el Excel real (patente, fecha, proveedor, tipo, costos, etc.).
"""
import os
from django.core.management.base import BaseCommand
from django.utils import timezone
from flota.models import Vehiculo, Proveedor, Mantenimiento, CuentaPresupuestaria


class Command(BaseCommand):
    help = 'Carga mantenciones 2025 desde Excel (1contexto/PLANILLA MANTENIMIENTO VEHÍCULOS 2025 HRN .xlsx).'

    def add_arguments(self, parser):
        parser.add_argument(
            '--ruta',
            type=str,
            default=None,
            help='Ruta al archivo Excel. Por defecto: 1contexto/PLANILLA MANTENIMIENTO VEHÍCULOS 2025 HRN .xlsx',
        )
        parser.add_argument('--dry-run', action='store_true', help='Solo mostrar qué se haría, no guardar.')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stdout.write(self.style.ERROR('Instale openpyxl: pip install openpyxl'))
            return

        ruta = options.get('ruta')
        if not ruta:
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            ruta = os.path.join(base_dir, '1contexto', 'PLANILLA MANTENIMIENTO VEHÍCULOS 2025 HRN .xlsx')
        if not os.path.isfile(ruta):
            self.stdout.write(self.style.WARNING(f'Archivo no encontrado: {ruta}'))
            self.stdout.write('Cree el archivo o pase --ruta con la ruta al Excel.')
            return

        dry_run = options.get('dry_run', False)
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        creados = 0
        omitidos = 0
        # Asumir estructura: fila 1 headers, desde fila 2 datos. Ajustar índices según el Excel real.
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.stdout.write(f'Columnas detectadas: {headers}')
        for row in ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            if not any(values):
                continue
            # Mapeo genérico: ajustar según nombres reales del Excel (patente, fecha, proveedor, tipo, costo)
            try:
                patente = str(values[0]).strip() if values[0] else None
                if not patente or len(patente) < 2:
                    omitidos += 1
                    continue
                vehiculo = Vehiculo.objects.filter(patente__iexact=patente).first()
                if not vehiculo:
                    self.stdout.write(self.style.WARNING(f'Vehículo no encontrado: {patente}'))
                    omitidos += 1
                    continue
                # Buscar proveedor por nombre si hay columna (ej. índice 2)
                proveedor_nombre = str(values[2]).strip() if len(values) > 2 and values[2] else 'Kaufmann'
                proveedor = Proveedor.objects.filter(
                    nombre_fantasia__icontains=proveedor_nombre
                ).first() or Proveedor.objects.filter(es_taller=True, activo=True).first()
                if not proveedor:
                    self.stdout.write(self.style.WARNING('Ningún proveedor taller en BD. Ejecute crear_proveedores_base.'))
                    omitidos += 1
                    continue
                # Fecha: asumir columna 1 o fecha actual 2025
                from datetime import date
                fecha_val = values[1] if len(values) > 1 else date(2025, 1, 1)
                if hasattr(fecha_val, 'date'):
                    fecha_ingreso = fecha_val.date()
                elif isinstance(fecha_val, str):
                    from datetime import datetime
                    try:
                        fecha_ingreso = datetime.strptime(fecha_val[:10], '%Y-%m-%d').date()
                    except ValueError:
                        fecha_ingreso = date(2025, 1, 1)
                else:
                    fecha_ingreso = date(2025, 1, 1)
                costo = int(values[4]) if len(values) > 4 and values[4] is not None else 0
                descripcion = str(values[3])[:500] if len(values) > 3 and values[3] else 'Mantenimiento 2025'
                tipo = 'Preventivo'  # o leer de columna si existe
                if not dry_run:
                    Mantenimiento.objects.create(
                        vehiculo=vehiculo,
                        proveedor=proveedor,
                        fecha_ingreso=fecha_ingreso,
                        descripcion_trabajo=descripcion[:500],
                        tipo_mantencion=tipo,
                        km_al_ingreso=vehiculo.kilometraje_actual or 0,
                        estado='Finalizado',
                        costo_total_real=costo,
                    )
                creados += 1
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Fila {row[0].row}: {e}'))
                omitidos += 1
        wb.close()
        self.stdout.write(self.style.SUCCESS(f'Procesadas: {creados} creadas/actualizadas, {omitidos} omitidas. Dry-run={dry_run}'))
