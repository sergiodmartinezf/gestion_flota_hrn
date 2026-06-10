from django import forms
from django.forms import formset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from ..models import (
    HojaRuta, CargaCombustible, FallaReportada, Alerta, Viaje, Vehiculo, Usuario,
    PacienteTraslado, PacienteViaje, DESTINOS_COMUNES,
)
from ..forms import HojaRutaForm, CargaCombustibleForm, FallaReportadaForm, ViajeForm, PacienteFormSet
from .utilidades import es_conductor_o_admin, es_conductor
from ..validators import normalizar_rut
from datetime import datetime, timedelta


TIPOS_SERVICIO = [
    ('Llamado', 'Llamado'),
    ('Rescate de Paciente', 'Rescate de Paciente'),
    ('A Urgencia HBO', 'A Urgencia HBO'),
    ('Exámenes', 'Exámenes'),
    ('Alta a Domicilio', 'Alta a Domicilio'),
    ('Interconsulta', 'Interconsulta'),
    ('Horas a especialista', 'Horas a especialista'),
    ('Imagen', 'Imagen'),
    ('Otro', 'Otro'),
]


@login_required
def acceso_bitacora(request):
    """
    Redirige según rol: admin/listado o conductor/hoja activa.
    """
    if request.user.rol in ['Administrador', 'Visualizador']:
        return redirect('listar_bitacoras')

    if request.user.rol == 'Conductor':
        # Buscar cualquier bitácora abierta del conductor, sin filtrar por fecha
        hoja_activa = HojaRuta.objects.filter(
            conductor=request.user,
            abierta=True
        ).order_by('-fecha', '-creado_en').first()

        if hoja_activa:
            return redirect('agregar_viaje', id=hoja_activa.id)
        else:
            return redirect('registrar_bitacora')

    return redirect('listar_bitacoras')


@login_required
def cerrar_hoja_ruta(request, id):
    hoja = get_object_or_404(HojaRuta, id=id)

    if request.user != hoja.conductor and request.user.rol != 'Administrador':
        messages.error(request, "No tienes permiso para cerrar esta hoja de ruta.")
        return redirect('agregar_viaje', id=hoja.id)

    if request.method == 'POST':
        try:
            km_final_input = request.POST.get('km_cierre')
            
            if not km_final_input:
                messages.error(request, "Debe ingresar el kilometraje final.")
                return redirect('agregar_viaje', id=hoja.id)
                
            km_final = int(km_final_input)

            if km_final < hoja.km_inicio:
                messages.error(request, f"El KM final ({km_final}) no puede ser menor al inicial ({hoja.km_inicio}).")
                return redirect('agregar_viaje', id=hoja.id)

            ultimo_viaje = hoja.viajes.order_by('-km_llegada').first()
            if ultimo_viaje and ultimo_viaje.km_llegada and km_final < ultimo_viaje.km_llegada:
                messages.warning(request, f"Atención: El KM de cierre es menor al del último viaje registrado ({ultimo_viaje.km_llegada}).")

            hoja.km_fin = km_final
            hoja.abierta = False
            hoja.save(update_fields=['km_fin', 'abierta'])

            vehiculo = hoja.vehiculo
            if km_final > vehiculo.kilometraje_actual:
                vehiculo.kilometraje_actual = km_final
                vehiculo.save()

            messages.success(request, f"Turno cerrado correctamente. KM Final: {km_final}. Total recorrido: {hoja.km_recorridos} km.")
            return redirect('listar_bitacoras')
            
        except ValueError:
            messages.error(request, "El kilometraje debe ser un número válido.")
            return redirect('agregar_viaje', id=hoja.id)
            
    return redirect('agregar_viaje', id=hoja.id)

@login_required
@user_passes_test(es_conductor_o_admin)
def registrar_bitacora(request):
    if request.method == 'POST':
        form = HojaRutaForm(request.POST)
        if form.is_valid():
            hoja_ruta = form.save(commit=False)
            hoja_ruta.conductor = request.user
            hoja_ruta.save()
            messages.success(request, 'Hoja de ruta creada exitosamente.')
            return redirect('agregar_viaje', id=hoja_ruta.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = HojaRutaForm()

    vehiculos = Vehiculo.queryset_para_hoja_ruta()
    vehiculos_info = [
        {
            'patente': v.patente,
            'texto': f"{v.patente} - {v.marca} {v.modelo} ({v.tipo_carroceria}) - {v.kilometraje_actual} km",
            'kilometraje': str(v.kilometraje_actual),   # ← convertir a string sin separadores
            'es_camioneta': v.tipo_carroceria == 'Camioneta'
        }
        for v in vehiculos
    ]
    return render(request, 'flota/registrar_hoja_ruta.html', {
        'form': form,
        'vehiculos': vehiculos,
        'vehiculos_info': vehiculos_info
    })

@login_required
@user_passes_test(es_conductor)
def agregar_viaje(request, id):
    hoja = get_object_or_404(HojaRuta, id=id)
    if not hoja.abierta:
        messages.warning(request, 'Esta hoja de ruta ya está cerrada. Debe iniciar una nueva hoja para registrar más viajes.')
        if request.user.rol == 'Conductor':
            return redirect('registrar_bitacora')
        return redirect('listar_bitacoras')

    vehiculo_tipo = hoja.vehiculo.tipo_carroceria
    
    ultimo_viaje = hoja.viajes.order_by('-km_llegada').first()
    km_sugerido = ultimo_viaje.km_llegada if ultimo_viaje and ultimo_viaje.km_llegada else hoja.km_inicio

    if request.method == 'POST':
        form = ViajeForm(request.POST, vehiculo_tipo=vehiculo_tipo)
        paciente_formset = PacienteFormSet(
            request.POST,
            request.FILES,
            form_kwargs={'vehiculo_tipo': vehiculo_tipo},
        )

        if form.is_valid():
            viaje = form.save(commit=False)
            viaje.hoja_ruta = hoja
            
            # Validaciones que dependen del objeto viaje
            km_actual_vehiculo = hoja.vehiculo.kilometraje_actual
            errores = False
            
            if viaje.km_salida < km_actual_vehiculo:
                form.add_error('km_salida', f'El KM de salida ({viaje.km_salida}) no puede ser menor al kilometraje actual del vehículo ({km_actual_vehiculo}).')
                errores = True
            elif viaje.km_salida < km_sugerido:
                form.add_error('km_salida', f'Error: KM salida ({viaje.km_salida}) es menor al anterior ({km_sugerido}).')
                errores = True
            
            if not errores and paciente_formset.is_valid():
                viaje.save()
                
                # Actualizar kilometraje del vehículo
                if viaje.km_llegada and viaje.km_llegada > hoja.vehiculo.kilometraje_actual:
                    hoja.vehiculo.kilometraje_actual = viaje.km_llegada
                    hoja.vehiculo.save()
                
                # Guardar pacientes
                pacientes = paciente_formset.save(commit=False)
                for paciente in pacientes:
                    paciente.viaje = viaje
                    if paciente.rut:
                        rut_norm = normalizar_rut(paciente.rut) or paciente.rut.strip()
                        paciente.rut = rut_norm
                        pv, _ = PacienteViaje.objects.get_or_create(rut=rut_norm)
                        paciente.paciente_viaje = pv
                    paciente.save()
                
                for form_del in paciente_formset.deleted_forms:
                    if form_del.instance.pk:
                        form_del.instance.delete()
                
                messages.success(request, 'Viaje registrado correctamente.')
                return redirect('agregar_viaje', id=hoja.id)
            else:
                if not errores:
                    messages.error(request, 'Error en los datos de los pacientes.')
        else:
            # Si el formulario principal no es válido, se mostrarán los errores en el template
            pass

    else:
        km_sugerido = ultimo_viaje.km_llegada if ultimo_viaje and ultimo_viaje.km_llegada else hoja.km_inicio
        initial_data = {
            'km_salida': km_sugerido,
            'km_llegada': km_sugerido if km_sugerido is not None else None,
        }
        form = ViajeForm(initial=initial_data, vehiculo_tipo=vehiculo_tipo)
        paciente_formset = PacienteFormSet(
            queryset=PacienteTraslado.objects.none(),
            form_kwargs={'vehiculo_tipo': vehiculo_tipo},
        )

    pacientes_anteriores = PacienteViaje.objects.all().order_by('rut')[:300]
    km_actual_vehiculo = hoja.vehiculo.kilometraje_actual
    if km_actual_vehiculo < hoja.km_inicio:
        km_actual_vehiculo = hoja.km_inicio
    
    context = {
        'hoja': hoja,
        'form': form,
        'paciente_formset': paciente_formset,
        'ultimos_viajes': hoja.viajes.all().order_by('-id')[:5],
        'pacientes_anteriores': pacientes_anteriores,
        'km_actual_vehiculo': km_actual_vehiculo,
        'destinos_choices': DESTINOS_COMUNES,
    }
    return render(request, 'flota/registrar_viaje.html', context)

@login_required
def listar_bitacoras(request):
    # Filtro base según rol de usuario
    if request.user.rol == 'Conductor':
        bitacoras = HojaRuta.objects.filter(conductor=request.user).order_by('-fecha', '-creado_en')
    else:
        bitacoras = HojaRuta.objects.all()

    # Filtros
    desde_filtro = request.GET.get('desde', '')
    hasta_filtro = request.GET.get('hasta', '')
    conductor_filtro = request.GET.get('conductor')
    estado_filtro = request.GET.get('estado')
    vehiculo_filtro = request.GET.get('vehiculo')

    if desde_filtro:
        bitacoras = bitacoras.filter(fecha__gte=desde_filtro)

    if hasta_filtro:
        bitacoras = bitacoras.filter(fecha__lte=hasta_filtro)

    if conductor_filtro:
        bitacoras = bitacoras.filter(conductor__rut=conductor_filtro)

    if estado_filtro:
        if estado_filtro == 'abierta':
            bitacoras = bitacoras.filter(abierta=True)
        elif estado_filtro == 'cerrada':
            bitacoras = bitacoras.filter(abierta=False)

    if vehiculo_filtro:
        bitacoras = bitacoras.filter(vehiculo__patente=vehiculo_filtro)

    bitacoras = bitacoras.order_by('-fecha', '-creado_en')

    # Obtener datos para filtros
    vehiculos = Vehiculo.objects.all().order_by('patente')
    conductores = Usuario.objects.filter(rol='Conductor', activo=True).order_by('nombre', 'apellido')

    return render(request, 'flota/listar_bitacoras.html', {
        'bitacoras': bitacoras,
        'vehiculos': vehiculos,
        'conductores': conductores,
        'vehiculo_filtro': vehiculo_filtro,
        'desde_filtro': desde_filtro,
        'hasta_filtro': hasta_filtro,
        'conductor_filtro': conductor_filtro,
        'estado_filtro': estado_filtro,
    })

@login_required
@user_passes_test(es_conductor_o_admin)
def modificar_bitacora(request, id):
    hoja = get_object_or_404(HojaRuta, id=id)

    if request.user != hoja.conductor and request.user.rol != 'Administrador':
        messages.error(request, "No tienes permiso para modificar esta hoja de ruta.")
        return redirect('detalle_bitacora', id=hoja.id)

    if request.method == 'POST':
        form = HojaRutaForm(request.POST, instance=hoja)
        if form.is_valid():
            form.save()
            messages.success(request, f'Hoja de ruta {hoja.id} actualizada exitosamente.')
            return redirect('detalle_bitacora', id=hoja.id)
    else:
        form = HojaRutaForm(instance=hoja)
    
    return render(request, 'flota/modificar_bitacora.html', {
        'form': form,
        'hoja': hoja
    })


@login_required
def detalle_bitacora(request, id):
    hoja = get_object_or_404(HojaRuta, id=id)
    viajes = hoja.viajes.prefetch_related('pacientes').all()

    total_km_viajes = sum(v.km_recorridos_calculados for v in viajes)

    if viajes.exists():
        km_llegadas = [v.km_llegada for v in viajes if v.km_llegada is not None]
        km_fin_hoja = max(km_llegadas) if km_llegadas else hoja.km_fin or hoja.km_inicio
    else:
        km_fin_hoja = hoja.km_fin or hoja.km_inicio

    return render(request, 'flota/detalle_bitacora.html', {
        'hoja': hoja,
        'viajes': viajes,
        'total_km_viajes': total_km_viajes,
        'km_fin_hoja': km_fin_hoja,
    })


@login_required
@user_passes_test(es_conductor_o_admin)
def registrar_carga_combustible(request):
    if request.method == 'POST':
        form = CargaCombustibleForm(request.POST)
        if form.is_valid():
            try:
                carga = form.save(commit=False)
                if request.user.rol == 'Conductor':
                    carga.conductor = request.user
                carga.save()
                # Actualizar kilometraje del vehículo
                if carga.kilometraje_al_cargar:
                    vehiculo = carga.patente_vehiculo
                    if carga.kilometraje_al_cargar > vehiculo.kilometraje_actual:
                        vehiculo.kilometraje_actual = carga.kilometraje_al_cargar
                        vehiculo.save()
                messages.success(request, 'Carga de combustible registrada exitosamente.')
                return redirect('listar_cargas_combustible')
            except Exception as e:
                messages.error(request, f'Error al registrar carga de combustible: {str(e)}')
        else:
            # Mostrar errores específicos
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CargaCombustibleForm()
    
    return render(request, 'flota/registrar_carga_combustible.html', {'form': form})


@login_required
def listar_cargas_combustible(request):
    # Filtro base según rol de usuario
    if request.user.rol == 'Conductor':
        cargas = CargaCombustible.objects.filter(conductor=request.user).order_by('-fecha')
    else:
        cargas = CargaCombustible.objects.all()

    # Filtros
    desde_filtro = request.GET.get('desde', '')
    hasta_filtro = request.GET.get('hasta', '')
    conductor_filtro = request.GET.get('conductor')
    vehiculo_filtro = request.GET.get('vehiculo')

    if desde_filtro:
        cargas = cargas.filter(fecha__gte=desde_filtro)

    if hasta_filtro:
        cargas = cargas.filter(fecha__lte=hasta_filtro)

    if conductor_filtro:
        cargas = cargas.filter(conductor__rut=conductor_filtro)

    if vehiculo_filtro:
        cargas = cargas.filter(patente_vehiculo__patente=vehiculo_filtro)

    cargas = cargas.order_by('-fecha')

    # Obtener datos para filtros
    vehiculos = Vehiculo.objects.all().order_by('patente')
    conductores = Usuario.objects.filter(rol='Conductor', activo=True).order_by('nombre', 'apellido')

    return render(request, 'flota/listar_cargas_combustible.html', {
        'cargas': cargas,
        'vehiculos': vehiculos,
        'conductores': conductores,
        'vehiculo_filtro': vehiculo_filtro,
        'desde_filtro': desde_filtro,
        'hasta_filtro': hasta_filtro,
        'conductor_filtro': conductor_filtro,
    })

@login_required
@user_passes_test(es_conductor_o_admin)
def registrar_incidente(request):
    if request.method == 'POST':
        form = FallaReportadaForm(request.POST)
        if form.is_valid():
            try:
                falla = form.save(commit=False)
                falla.conductor = request.user
                falla.save()

                # Crear alerta por falla cuando se supera el umbral de mantención
                vehiculo = falla.vehiculo
                if vehiculo.umbral_mantencion > 0 and vehiculo.kilometraje_actual >= vehiculo.umbral_mantencion:
                    Alerta.objects.create(
                        vehiculo=vehiculo,
                        descripcion=f'Falla reportada: {falla.descripcion}',
                        valor_umbral=vehiculo.kilometraje_actual,
                    )
                
                messages.success(request, 'Incidente registrado exitosamente.')
                return redirect('listar_incidentes')
            except Exception as e:
                messages.error(request, f'Error al registrar incidente: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FallaReportadaForm()
        # Establecer fecha actual por defecto
        if not form.instance.pk:
            from datetime import date
            form.fields['fecha_reporte'].initial = date.today()
    
    return render(request, 'flota/registrar_incidente.html', {'form': form})


@login_required
def listar_incidentes(request):
    # Filtro base según rol de usuario
    if request.user.rol == 'Conductor':
        incidentes = FallaReportada.objects.filter(conductor=request.user).order_by('-fecha_reporte')
    else:
        incidentes = FallaReportada.objects.all()

    # Filtros
    desde_filtro = request.GET.get('desde', '')
    hasta_filtro = request.GET.get('hasta', '')
    conductor_filtro = request.GET.get('conductor')
    vehiculo_filtro = request.GET.get('vehiculo')

    if desde_filtro:
        incidentes = incidentes.filter(fecha_reporte__gte=desde_filtro)

    if hasta_filtro:
        incidentes = incidentes.filter(fecha_reporte__lte=hasta_filtro)

    if conductor_filtro:
        incidentes = incidentes.filter(conductor__rut=conductor_filtro)

    if vehiculo_filtro:
        incidentes = incidentes.filter(vehiculo__patente=vehiculo_filtro)

    incidentes = incidentes.order_by('-fecha_reporte')

    # Obtener datos para filtros
    vehiculos = Vehiculo.objects.all().order_by('patente')
    conductores = Usuario.objects.filter(rol='Conductor', activo=True).order_by('nombre', 'apellido')

    return render(request, 'flota/listar_incidentes.html', {
        'incidentes': incidentes,
        'vehiculos': vehiculos,
        'conductores': conductores,
        'vehiculo_filtro': vehiculo_filtro,
        'desde_filtro': desde_filtro,
        'hasta_filtro': hasta_filtro,
        'conductor_filtro': conductor_filtro,
    })


@login_required
@user_passes_test(es_conductor)
def historial_conductor(request):
    # Base querysets filtrados por conductor
    bitacoras_qs = HojaRuta.objects.filter(conductor=request.user).order_by('-fecha', '-creado_en')
    cargas_qs = CargaCombustible.objects.filter(conductor=request.user).order_by('-fecha')
    incidentes_qs = FallaReportada.objects.filter(conductor=request.user).order_by('-fecha_reporte')

    # Filtros para Bitácoras (prefijo b_)
    b_desde = request.GET.get('b_desde')
    b_hasta = request.GET.get('b_hasta')
    b_vehiculo = request.GET.get('b_vehiculo')
    if b_desde:
        bitacoras_qs = bitacoras_qs.filter(fecha__gte=b_desde)
    if b_hasta:
        bitacoras_qs = bitacoras_qs.filter(fecha__lte=b_hasta)
    if b_vehiculo:
        bitacoras_qs = bitacoras_qs.filter(vehiculo__patente=b_vehiculo)

    # Filtros para Cargas (prefijo c_)
    c_desde = request.GET.get('c_desde')
    c_hasta = request.GET.get('c_hasta')
    c_vehiculo = request.GET.get('c_vehiculo')
    if c_desde:
        cargas_qs = cargas_qs.filter(fecha__gte=c_desde)
    if c_hasta:
        cargas_qs = cargas_qs.filter(fecha__lte=c_hasta)
    if c_vehiculo:
        cargas_qs = cargas_qs.filter(patente_vehiculo__patente=c_vehiculo)

    # Filtros para Incidentes (prefijo i_)
    i_desde = request.GET.get('i_desde')
    i_hasta = request.GET.get('i_hasta')
    i_vehiculo = request.GET.get('i_vehiculo')
    if i_desde:
        incidentes_qs = incidentes_qs.filter(fecha_reporte__gte=i_desde)
    if i_hasta:
        incidentes_qs = incidentes_qs.filter(fecha_reporte__lte=i_hasta)
    if i_vehiculo:
        incidentes_qs = incidentes_qs.filter(vehiculo__patente=i_vehiculo)

    # Obtener IDs de vehículos únicos de los tres querysets (sin usar union().distinct())
    vehiculos_ids = set()
    vehiculos_ids.update(bitacoras_qs.values_list('vehiculo_id', flat=True))
    vehiculos_ids.update(cargas_qs.values_list('patente_vehiculo_id', flat=True))
    vehiculos_ids.update(incidentes_qs.values_list('vehiculo_id', flat=True))
    vehiculos_usados = Vehiculo.objects.filter(id__in=vehiculos_ids).order_by('patente')

    context = {
        'bitacoras': bitacoras_qs[:50],
        'cargas': cargas_qs[:50],
        'incidentes': incidentes_qs[:50],
        'vehiculos': vehiculos_usados,
        'b_desde': b_desde,
        'b_hasta': b_hasta,
        'b_vehiculo': b_vehiculo,
        'c_desde': c_desde,
        'c_hasta': c_hasta,
        'c_vehiculo': c_vehiculo,
        'i_desde': i_desde,
        'i_hasta': i_hasta,
        'i_vehiculo': i_vehiculo,
    }
    return render(request, 'flota/historial_conductor.html', context)


@login_required
def reabrir_bitacora(request, id):
    hoja = get_object_or_404(HojaRuta, id=id)

    # Verificar permisos: el conductor de la hoja o administrador
    if request.user != hoja.conductor and request.user.rol != 'Administrador':
        messages.error(request, "No tienes permiso para reabrir esta hoja de ruta.")
        return redirect('historial_conductor')

    if hoja.abierta:
        messages.warning(request, "La hoja de ruta ya está abierta.")
        return redirect('agregar_viaje', id=hoja.id)

    # Verificar que no exista otra bitácora abierta para este conductor
    otra_abierta = HojaRuta.objects.filter(
        conductor=request.user,
        abierta=True
    ).exclude(id=hoja.id).exists()

    if otra_abierta:
        messages.error(
            request,
            "No se puede reabrir esta bitácora porque ya tienes otra hoja de ruta abierta. "
            "Por favor, cierra la bitácora activa primero."
        )
        return redirect('historial_conductor')

    # Reabrir: establecer abierta=True y eliminar km_fin
    hoja.abierta = True
    hoja.km_fin = None
    hoja.save(update_fields=['abierta', 'km_fin'])

    messages.success(
        request,
        f"Hoja de ruta del {hoja.fecha.strftime('%d/%m/%Y')} reabierta correctamente. "
        "Puedes agregar más viajes."
    )
    return redirect('agregar_viaje', id=hoja.id)


@login_required
def exportar_traslados_form(request):
    """
    Muestra formulario con filtros (desde, hasta, vehículo, conductor) para exportar Excel.
    """
    vehiculos = Vehiculo.objects.all().order_by('patente')
    conductores = Usuario.objects.filter(rol='Conductor', activo=True).order_by('nombre', 'apellido')
    return render(request, 'flota/exportar_traslados.html', {
        'vehiculos': vehiculos,
        'conductores': conductores,
    })


@login_required
def exportar_consolidado_viajes(request):
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    vehiculo_filtro = request.GET.get('vehiculo')
    conductor_filtro = request.GET.get('conductor')

    viajes = Viaje.objects.select_related(
        'hoja_ruta', 'hoja_ruta__vehiculo', 'hoja_ruta__conductor'
    ).prefetch_related('pacientes').all()

    if desde:
        try:
            fecha_desde = datetime.strptime(desde, '%Y-%m-%d').date()
            viajes = viajes.filter(hoja_ruta__fecha__gte=fecha_desde)
        except ValueError:
            pass
    if hasta:
        try:
            fecha_hasta = datetime.strptime(hasta, '%Y-%m-%d').date()
            viajes = viajes.filter(hoja_ruta__fecha__lte=fecha_hasta)
        except ValueError:
            pass
    if vehiculo_filtro:
        viajes = viajes.filter(hoja_ruta__vehiculo__patente=vehiculo_filtro)
    if conductor_filtro:
        viajes = viajes.filter(hoja_ruta__conductor__rut=conductor_filtro)
    viajes = viajes.order_by('-hoja_ruta__fecha')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Traslados'

    headers = [
        'Fecha', 'Vehículo', 'Conductor', 'Turno', 'Hora Salida', 'Hora Llegada',
        'Destino', 'RUT Paciente', 'Categoría Traslado', 'Sentido', 'Kms Viaje'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    row_num = 2
    for viaje in viajes:
        hoja = viaje.hoja_ruta
        fecha_str = hoja.fecha.strftime('%d-%m-%Y')
        vehiculo_patente = hoja.vehiculo.patente
        conductor_nombre = hoja.conductor.nombre_completo
        turno = hoja.turno
        hora_salida = viaje.hora_salida.strftime('%H:%M')
        hora_llegada = viaje.hora_llegada.strftime('%H:%M') if viaje.hora_llegada else '-'
        km_viaje = viaje.km_recorridos_calculados

        pacientes = list(viaje.pacientes.all())
        if not pacientes:
            ws.cell(row=row_num, column=1, value=fecha_str)
            ws.cell(row=row_num, column=2, value=vehiculo_patente)
            ws.cell(row=row_num, column=3, value=conductor_nombre)
            ws.cell(row=row_num, column=4, value=turno)
            ws.cell(row=row_num, column=5, value=hora_salida)
            ws.cell(row=row_num, column=6, value=hora_llegada)
            ws.cell(row=row_num, column=7, value='-')
            ws.cell(row=row_num, column=8, value='Sin pacientes')
            ws.cell(row=row_num, column=9, value='-')
            ws.cell(row=row_num, column=10, value='-')
            ws.cell(row=row_num, column=11, value=km_viaje)
            row_num += 1
        else:
            for p in pacientes:
                destino_display = p.get_destino_tipo_display()
                if p.direccion_especifica and p.destino_tipo == 'DOMICILIO':
                    destino_display += f" - {p.direccion_especifica}"
                ws.cell(row=row_num, column=1, value=fecha_str)
                ws.cell(row=row_num, column=2, value=vehiculo_patente)
                ws.cell(row=row_num, column=3, value=conductor_nombre)
                ws.cell(row=row_num, column=4, value=turno)
                ws.cell(row=row_num, column=5, value=hora_salida)
                ws.cell(row=row_num, column=6, value=hora_llegada)
                ws.cell(row=row_num, column=7, value=destino_display)
                ws.cell(row=row_num, column=8, value=p.rut or '')
                ws.cell(row=row_num, column=9, value=p.get_categoria_traslado_display())
                ws.cell(row=row_num, column=10, value=p.get_sentido_display())
                ws.cell(row=row_num, column=11, value=km_viaje)
                row_num += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="consolidado_traslados_{timezone.now().date()}.xlsx"'
    )
    wb.save(response)
    return response
    